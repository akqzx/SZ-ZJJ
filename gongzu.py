import os
import time
import datetime
import openpyxl
import sys
import platform
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.chart.axis import DateAxis
from selenium.webdriver.common.by import By
from openpyxl.chart import (
    LineChart,
    Reference,
)
gongzubeianhao = ""
xingming = ""
shenfenID = ""

platform_ = platform.system()
if platform_ == "Windows":
    driver = webdriver.Chrome(ChromeDriverManager().install())
elif platform_ == "Linux":
    from selenium.webdriver.chrome.options import Options
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--no-sandbox')
    driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
elif platform_ == "Mac":
    print("not support")
    sys.exit()
if not os.path.exists("gongzu.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws['A1'] = '时间'
    ws['B1'] = '公租户籍区排序号'
    ws['C1'] = '公租轮候排序号'
    wb.save('gongzu.xlsx')
###公租
driver.get('https://zjj.sz.gov.cn/zfxx/bzflh/?path=main/#/lhmcgr')
time.sleep(1)
driver.find_element(By.XPATH, '//*[@id="app"]/section/header/form/div[1]/div/div/div/input').click()
time.sleep(1)
driver.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[1]/ul/li[2]').click()
driver.find_element(By.XPATH, '//*[@id="app"]/section/header/form/div[2]/div/div/input').send_keys(gongzubeianhao)
driver.find_element(By.XPATH, '//*[@id="app"]/section/header/form/div[3]/div/div/input').send_keys(xingming)
driver.find_element(By.XPATH, '//*[@id="app"]/section/header/form/div[4]/div/div/input').send_keys(shenfenID)
driver.find_element(By.XPATH, '//*[@id="app"]/section/header/form/div[5]/div/button[1]/span').click()
time.sleep(3)
gongzuqu = driver.find_element(By.XPATH, '//*[@id="app"]/section/main/div/div[3]/table/tbody/tr/td[2]').text
gongzushi = driver.find_element(By.XPATH, '//*[@id="app"]/section/main/div/div[3]/table/tbody/tr/td[3]').text

###查询上一次排位信息
wb = openpyxl.load_workbook('gongzu.xlsx')
ws = wb['Sheet1']
oldgongzuqudyg = ("B"+str(ws.max_row))
oldgongzuqu = ws[oldgongzuqudyg].value

###判断是否与上一次数值相同
if int(gongzuqu) == oldgongzuqu:
    print("与上次查询的数据一致，结束")
    sys.exit(0)

line = str(ws.max_row+1)
todyA = ("A"+line)
gongzuqub = ("B"+line)
gongzushic = ("C"+line)
ws[todyA] = datetime.datetime.now().strftime("%Y-%m-%d")
ws[gongzuqub] = float(gongzuqu)
ws[gongzushic] = float(gongzushi)
copy_sheet1 = wb.copy_worksheet(wb.worksheets[0])
copy_sheet1.title = "Sheet2"
del wb["Sheet1"]
ws = wb["Sheet2"]
ws.title = 'Sheet1'
ws.column_dimensions['A'].width = 10.5
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 14

data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=line)
c1 = LineChart()
c1.title = "公租户籍区排序号"
c1.style = 14
c1.y_axis.crossAx = 500
c1.x_axis = DateAxis(crossAx=100)
c1.x_axis.number_format = 'd-mmm'
c1.x_axis.majorTimeUnit = "days"
c1.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=line)
c1.set_categories(dates)
ws.add_chart(c1, "E3")

data2 = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=line)
c2 = LineChart()
c2.title = "公租轮候排序号"
c2.style = 12
c2.y_axis.crossAx = 500
c2.x_axis = DateAxis(crossAx=100)
c2.x_axis.number_format = 'd-mmm'
c2.x_axis.majorTimeUnit = "days"
c2.add_data(data2, titles_from_data=True)
dates2 = Reference(ws, min_col=1, min_row=2, max_row=line)
c2.set_categories(dates2)
ws.add_chart(c2, "E20")

wb.save('gongzu.xlsx')
driver.close()
driver.quit()