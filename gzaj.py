import os
import sys
import time
import datetime
import platform
import openpyxl
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.chart.axis import DateAxis
from openpyxl.chart import (
    LineChart,
    Reference,
)
gzbeianhao = "公租备案回执号"
xingming = "姓名"
shenfenID = "身份证号码"

platform_ = platform.system()
if platform_ == "Windows" or platform_ == "Darwin":
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver1 = webdriver.Chrome(ChromeDriverManager().install())
elif platform_ == "Linux":
    from selenium.webdriver.chrome.options import Options
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--no-sandbox')
    driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
    driver1 = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
else:
    print("未支持的操作系统")
    sys.exit()
if not os.path.exists("zjj.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws['A1'] = '时间'
    ws['B1'] = '公租区排位'
    ws['C1'] = '公租市排位'
    ws['D1'] = '安居市排位'
    ws['E1'] = '安居人口排位'
    wb.save('zjj.xlsx')
driver.get('http://zjj.sz.gov.cn/bzflh//jsplib/lhccx/singlelhc_query1.jsp')
driver.find_element_by_xpath('//*[@id="bahzh"]').send_keys(gzbeianhao)
driver.find_element_by_xpath('//*[@id="xingm"]').send_keys(xingming)
driver.find_element_by_xpath('//*[@id="sfz"]').send_keys(shenfenID)
driver.find_element_by_xpath('//*[@id="button"]').click()
time.sleep(1)
gzqpw = driver.find_element_by_xpath("//*[@id='table']/tbody/tr/td[2]").text
gzspw = driver.find_element_by_xpath("//*[@id='table']/tbody/tr/td[3]").text
driver1.get('http://zjj.sz.gov.cn/bzflh//lhmcAction.do?method=queryYgbLhmcInfo&waittype=1')
driver1.find_element_by_xpath('//*[@id="xingm"]').send_keys(xingming)
time.sleep(1)
driver1.find_element_by_xpath('//*[@id="button"]').click()
time.sleep(1)
ajlh = driver1.find_element_by_xpath("//*[@id='cusTable']/tbody/tr/td[1]").text
ajrk = driver1.find_element_by_xpath("//*[@id='cusTable']/tbody/tr/td[7]").text
wb = openpyxl.load_workbook('zjj.xlsx')
ws = wb['Sheet1']
line = str(ws.max_row+1)
wsa = ("A"+line)
wsb = ("B"+line)
wsc = ("C"+line)
wsd = ("D"+line)
wse = ("E"+line)
ws[wsa] = datetime.datetime.now().strftime("%Y-%m-%d")
ws[wsb] = float(gzqpw)
ws[wsc] = float(gzspw)
ws[wsd] = float(ajlh)
ws[wse] = float(ajrk)
copy_sheet1 = wb.copy_worksheet(wb.worksheets[0])
copy_sheet1.title = "Sheet2"
del wb["Sheet1"]
ws = wb["Sheet2"]
ws.title = 'Sheet1'
ws.column_dimensions['A'].width = 10.5
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 9.5
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 11
data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=line)
c1 = LineChart()
c1.title = "公租区排位"
c1.style = 14
c1.y_axis.crossAx = 500
c1.x_axis = DateAxis(crossAx=100)
c1.x_axis.number_format = 'd-mmm'
c1.x_axis.majorTimeUnit = "days"
c1.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=line)
c1.set_categories(dates)
ws.add_chart(c1, "G3")
data2 = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=line)
c2 = LineChart()
c2.title = "公租市排位"
c2.style = 12
c2.y_axis.crossAx = 500
c2.x_axis = DateAxis(crossAx=100)
c2.x_axis.number_format = 'd-mmm'
c2.x_axis.majorTimeUnit = "days"
c2.add_data(data2, titles_from_data=True)
dates2 = Reference(ws, min_col=1, min_row=2, max_row=line)
c2.set_categories(dates2)
ws.add_chart(c2, "G20")
data3 = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=line)
c3 = LineChart()
c3.title = "安居市排位"
c3.style = 11
c3.y_axis.crossAx = 500
c3.x_axis = DateAxis(crossAx=100)
c3.x_axis.number_format = 'd-mmm'
c3.x_axis.majorTimeUnit = "days"
c3.add_data(data3, titles_from_data=True)
dates3 = Reference(ws, min_col=1, min_row=2, max_row=line)
c3.set_categories(dates3)
ws.add_chart(c3, "O3")
data4 = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=line)
c4 = LineChart()
c4.title = "安居人口排位"
c4.style = 15
c4.y_axis.crossAx = 500
c4.x_axis = DateAxis(crossAx=100)
c4.x_axis.number_format = 'd-mmm'
c4.x_axis.majorTimeUnit = "days"
c4.add_data(data4, titles_from_data=True)
dates4 = Reference(ws, min_col=1, min_row=2, max_row=line)
c4.set_categories(dates4)
ws.add_chart(c4, "O20")
wb.save('zjj.xlsx')
driver.close()
driver.quit()
driver1.close()
driver1.quit()
