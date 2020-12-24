import os
import time
import datetime
import sys
import platform
import openpyxl
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.chart.axis import DateAxis
from openpyxl.chart import (
    LineChart,
    Reference,
)
beianhao = "备案回执号"
xingming = "姓名"
shenfenID = "身份证号码"

platform_ = platform.system()
if platform_ == "Windows":
    driver = webdriver.Chrome(ChromeDriverManager().install())
elif platform_ == "Linux":
    from selenium.webdriver.chrome.options import Options
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--no-sandbox')
    driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options)
elif platform_ == "Mac":
    print("not support")
    sys.exit()
if not os.path.exists("zjj.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws['A1'] = '时间'
    ws['B1'] = '户籍区排序号'
    ws['C1'] = '轮候排序号'
    wb.save('zjj.xlsx')
driver.get('http://zjj.sz.gov.cn/bzflh//jsplib/lhccx/singlelhc_query1.jsp')
driver.find_element_by_xpath('//*[@id="bahzh"]').send_keys(beianhao)
driver.find_element_by_xpath('//*[@id="xingm"]').send_keys(xingming)
driver.find_element_by_xpath('//*[@id="sfz"]').send_keys(shenfenID)
driver.find_element_by_xpath('//*[@id="button"]').click()
time.sleep(3)
qu = driver.find_element_by_xpath("//*[@id='table']/tbody/tr/td[2]").text
shi = driver.find_element_by_xpath("//*[@id='table']/tbody/tr/td[3]").text
wb = openpyxl.load_workbook('zjj.xlsx')
ws = wb['Sheet1']
line = str(ws.max_row+1)
todyA = ("A"+line)
qub = ("B"+line)
shic = ("C"+line)
ws[todyA] = datetime.datetime.now().strftime("%Y-%m-%d")
ws[qub] = float(qu)
ws[shic] = float(shi)
copy_sheet1 = wb.copy_worksheet(wb.worksheets[0])
copy_sheet1.title = "Sheet2"
del wb["Sheet1"]
ws = wb["Sheet2"]
ws.title = 'Sheet1'
ws.column_dimensions['A'].width = 10.5
ws.column_dimensions['B'].width = 12.5
ws.column_dimensions['C'].width = 9.5
data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=line)
c1 = LineChart()
c1.title = "户籍区排序号"
c1.style = 14
c1.y_axis.crossAx = 500
c1.x_axis = DateAxis(crossAx=100)
c1.x_axis.number_format = 'd-mmm'
c1.x_axis.majorTimeUnit = "days"
c1.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=1, min_row=2, max_row=line)
c1.set_categories(dates)
ws.add_chart(c1, "E3")
data2 = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=line)
c2 = LineChart()
c2.title = "轮候排序号"
c2.style = 12
c2.y_axis.crossAx = 500
c2.x_axis = DateAxis(crossAx=100)
c2.x_axis.number_format = 'd-mmm'
c2.x_axis.majorTimeUnit = "days"
c2.add_data(data2, titles_from_data=True)
dates2 = Reference(ws, min_col=1, min_row=2, max_row=line)
c2.set_categories(dates2)
ws.add_chart(c2, "E20")
wb.save('zjj.xlsx')
driver.close()
driver.quit()
